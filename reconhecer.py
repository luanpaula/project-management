import requests
import json
from openpyxl import load_workbook

baixados = 0
erros = 0

# Carrega cookies salvos manualmente do Chrome
with open('cookies.json', 'r') as f:
    raw_cookies = json.load(f)

cookies = {cookie['name']: cookie['value'] for cookie in raw_cookies}

# Lê a planilha
wb = load_workbook("data.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    nome_arquivo, url = row[0], row[1]

    try:
        response = requests.get(url, cookies=cookies)

        if response.status_code == 200:
            with open(f"{nome_arquivo}", 'wb') as f:
                f.write(response.content)
            print(f"[OK] Baixado: {nome_arquivo}")
            baixados += 1
        else:
            print(f"[Erro {response.status_code}] {url}")
            erros += 1

    except Exception as e:
        print(f"[Falha] {url}: {e}")

if(baixados != 0):
    print(f"🥳 {baixados} arquivos baixados")
if(erros != 0):
    print(f"😖 {erros} arquivos com erro")